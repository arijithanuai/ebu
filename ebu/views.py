
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Province, Kabupaten, Link, User, DrpFile, DBfile
from .forms import UserForm
import csv
import sys
from django.contrib import messages
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
import io
from shapely import wkt
import json
from django.contrib.gis.geos import GEOSGeometry
from django.urls import reverse
import base64
from django.shortcuts import render, redirect
from django.http import JsonResponse
from .models import Province, Kabupaten, Link, User, DrpFile, DBfile, Alignment
from .forms import UserForm
import csv
import sys
from django.contrib import messages
import pandas as pd
from django.views.decorators.csrf import csrf_exempt
import io
from shapely import wkt
import os, time ,json
import tempfile
import shutil
from django.contrib.gis.geos import GEOSGeometry
from django.urls import reverse
import base64
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from django.conf import settings
from .Scripts.main import runValidationScript

import json
import base64
from django.http import JsonResponse, HttpResponseForbidden
from django.shortcuts import render
from django.contrib.auth.decorators import login_required

# Decode and verify token
def decode_token(token):
    try:
        # Decode the base64 token
        decoded_data = base64.b64decode(token).decode('utf-8')
        # Parse the JSON data from the decoded token
        token_data = json.loads(decoded_data)
        return token_data
    except Exception as e:
        print(f"Error decoding token: {e}")
        return None


def location_selector(request):
    # ---------------- Token Handling ----------------
    token = request.GET.get('token', None)
    print(decode_token(token))
    if not token:
        return render(request, '403.html', status=403)

    # Decode the token
    token_data = decode_token(token)
    if not token_data:
        return HttpResponseForbidden("Invalid token")
    
# {'adminCode': '09-52-00', 'userRole': 'province_lg', 'userName': 'Prabin', 'userphone': '+628596741230', 'userEmail': 'prabin@gmail.com'}
    admin_code = token_data.get('adminCode')  # e.g. "IDN-52-00"
    user_role = token_data.get('userRole')    # "province_lg" / "kabupaten_lg"
    lg_user_name = token_data.get('userName')
    lg_email = token_data.get('userEmail')
    lg_ph_no = token_data.get('userphone')
                              
    # Split adminCode → extract province + kabupaten codes
    parts = admin_code.split("-")
    pCode = parts[1]  # province code, e.g. "52"
    kCode = parts[2]  # kabupaten code, e.g. "00"

    # Map role → status
    if user_role == "province_lg":
        status = "province"
    elif user_role == "kabupaten_lg":
        status = "kabupaten"
    else:
        status = ""

    # ---------------- Preselect Province & Kabupaten ----------------
    provinces = Province.objects.all()

    province_obj = Province.objects.filter(pCode=pCode).first()
    preselect_province = province_obj.id if province_obj else ""

    preselect_kabupaten = ""
    if status == "kabupaten" and province_obj:
        kab_obj = Kabupaten.objects.filter(province=province_obj, kCode=kCode).first()
        preselect_kabupaten = kab_obj.id if kab_obj else ""

    # ---------------- Handle POST (form submit) ----------------
    if request.method == 'POST':
        admcode = request.POST.get('admcode')
        lgName = request.POST.get('lgName')
        emailId = request.POST.get('emailId')
        phoneNumber = request.POST.get('phoneNumber')

        # Save user info
        user_obj = User.objects.create(
            admcode=admcode,
            lgName=lgName,
            emailId=emailId,
            phoneNumber=phoneNumber
        )

        # ----- Save Excel Data -----
        excel_linkcodes = set()
        file_content = request.session.get('validated_file')
        if file_content:
            file_stream = io.BytesIO(file_content.encode('latin1'))
            df = pd.read_excel(file_stream)

            excel_linkcodes = set(df["Link_Code"].astype(str))

            links_to_create = [
                Link(
                    admCode=row["Adm_Code"],
                    linkNo=row["Link_No"],
                    linkCode=row["Link_Code"],
                    linkName=row["Link_Name"],
                    linkLengthOfficial=row["Link_Length_Official"],
                    linkLengthActual=row["Link_Length_Actual"],
                    status=row["Status"]
                )
                for _, row in df.iterrows()
            ]
            Link.objects.bulk_create(links_to_create, ignore_conflicts=True)
            del request.session['validated_file']
            messages.success(request, f"✅ Link data uploaded successfully! {len(links_to_create)} records saved.")

        # ----- Save TXT Data (alignment) -----
        txt_content = request.session.get("validated_txt")
        if txt_content:
            txt_link_data = json.loads(txt_content)
            alignments_to_create = []

            # Mapping linkCode → Link object
            link_map = {l.linkCode: l for l in Link.objects.filter(linkCode__in=excel_linkcodes)}

            for linkcode, line_wkt in txt_link_data:
                if linkcode in link_map:
                    geom = GEOSGeometry(line_wkt, srid=4326)
                    alignments_to_create.append(
                        Alignment(admCode=admcode, linkNo=link_map[linkcode], linkGeometry=geom)
                    )

            if alignments_to_create:
                Alignment.objects.bulk_create(alignments_to_create, ignore_conflicts=True)
                messages.success(request, f"✅ Alignment data uploaded successfully! {len(alignments_to_create)} records saved.")
            else:
                messages.warning(request, "⚠ No matching LinkCode found between TXT and Excel data.")

            del request.session['validated_txt']

        # ----- Save DRP File -----
        if request.FILES.get("link_drpexcel"):
            drp_file = request.FILES["link_drpexcel"]
            DrpFile.objects.create(admCode=admcode, drpFile=drp_file)
            messages.success(request, f"✅ DRP file uploaded successfully: {drp_file.name}")

        # ----- Save DB File -----
        if request.FILES.get("db_file"):
            db_file = request.FILES["db_file"]
            DBfile.objects.create(admCode=admcode, fileUrl=db_file)
            messages.success(request, f"✅ DB file uploaded successfully: {db_file.name}")

        return redirect('done')
    print(f"-----------{lg_user_name}, {lg_email}, {lg_ph_no}")
    # ---------------- GET request → Render form ----------------
    return render(request, 'pk.html', {
        'provinces': provinces,
        "preselect_status": status,
        "preselect_province": preselect_province,
        "preselect_kabupaten": preselect_kabupaten,
        "lg_user_name" : lg_user_name,
        "lg_email" : lg_email,
        "lg_ph_no" :lg_ph_no,
    })


def data_updated(request):
    return render(request, "done.html")

def get_kabupatens(request):
    province_id = request.GET.get('province_id')
    kabupatens = Kabupaten.objects.filter(province_id=province_id).values('id', 'admNameEng', 'kCode')
    return JsonResponse(list(kabupatens), safe=False)

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import base64
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import base64

@csrf_exempt
def validate_link_excel(request):
    if request.method == "POST" and request.FILES.get("link_excel"):
        file = request.FILES["link_excel"]
        admcode_from_form = request.POST.get("admcode", "").strip()
        
        # --- Extension validation ---
        if not file.name.lower().endswith(".xlsx"):
            return JsonResponse({
                "valid": False,
                "message": "❌ Invalid file type. Please upload an Excel file with .xlsx extension only."
            })

        if not admcode_from_form:
            return JsonResponse({
                "valid": False,
                "message": " Please select Status/Province/Kabupaten first to generate AdmCode before uploading Excel."
            })

        try:
            df = pd.read_excel(file)
        except Exception as e:
            return JsonResponse({"valid": False, "message": f" Invalid Excel file: {e}"})

        required_cols = [
            "Adm_Code", "Link_No", "Link_Code", "Link_Name",
            "Link_Length_Official", "Link_Length_Actual", "Status"
        ]
        errors = []
        row_error_map = {}

        # --- Schema validation ---
        missing = [col for col in required_cols if col not in df.columns]
        extra = [col for col in df.columns if col not in required_cols]

        if missing or extra:
            msg = []
            if missing:
                msg.append(f" Missing/Invalid columns: {', '.join(missing)}")
            if extra:
                msg.append(f" Extra/Unexpected columns: {', '.join(extra)}")
            return JsonResponse({
                "valid": False,
                "message": " Excel schema invalid ❌<br>" + "<br>".join(msg),
                "template_excel_url": reverse("download_template_excel")
            })

        if df.empty:
            return JsonResponse({"valid": False, "message": " Excel file contains no data"})

        # --- AdmCode validation ---
        unique_admcodes = df["Adm_Code"].dropna().unique()
        if len(unique_admcodes) > 1:
            errors.append(" Excel file contains multiple different Adm_Code values.")
        else:
            excel_admcode = str(unique_admcodes[0]).strip()
            if excel_admcode != admcode_from_form:
                errors.append(f" Adm_Code in Excel ({excel_admcode}) does not match selected AdmCode ({admcode_from_form}).")

        # --- Row-level validation ---
        seen_linknos, seen_linkcodes = set(), set()
        for idx, row in df.iterrows():
            row_errors = []
            link_code = str(row.get("Link_Code", "Unknown"))

            # Missing values
            missing_vals = [col for col in required_cols if pd.isnull(row[col])]
            if missing_vals:
                row_errors.append(f"Missing {', '.join(missing_vals)}")

            # Link_No must be integer
            try:
                int(row["Link_No"])
            except Exception:
                row_errors.append("Link_No must be an integer")

            # Status must be B/P/K
            if str(row["Status"]).strip().upper() not in ["B", "P", "K"]:
                row_errors.append("Status must be one of B, P, K")

            # Length numeric
            for col in ["Link_Length_Official", "Link_Length_Actual"]:
                try:
                    float(row[col])
                except Exception:
                    row_errors.append(f"{col} must be numeric")

            # Duplicates
            link_no = str(row["Link_No"])
            if link_no in seen_linknos:
                row_errors.append(f"Duplicate Link_No {link_no}")
            else:
                seen_linknos.add(link_no)

            if link_code in seen_linkcodes:
                row_errors.append(f"Duplicate Link_Code {link_code}")
            else:
                seen_linkcodes.add(link_code)

            if row_errors:
                row_error_map[idx] = row_errors

        # --- If errors exist → build error Excel ---
        if errors or row_error_map:
            file.seek(0)
            wb = load_workbook(file)
            ws = wb.active

            # Ensure H col header is "Error_Notes"
            ws["H1"].value = "Error_Notes"

            red_fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")

            for idx, row in df.iterrows():
                excel_row = idx + 2
                if idx in row_error_map:
                    # highlight full row
                    for col_idx in range(1, 8):  # A–G
                        ws.cell(row=excel_row, column=col_idx).fill = red_fill
                    # write errors in H
                    ws[f"H{excel_row}"].value = "; ".join(row_error_map[idx])

            # Save to memory
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Save as base64 in session
            request.session["error_excel"] = base64.b64encode(output.getvalue()).decode("utf-8")

            # Return row-wise errors also for frontend
            row_msgs = [f"❌ Row {i+2} (Link_Code {df.loc[i,'Link_Code']}): {', '.join(errs)}"
                        for i, errs in row_error_map.items()]

            return JsonResponse({
                "valid": False,
                "message": "Excel validation failed ❌<br>" + "<br>".join(errors + row_msgs),
                "error_excel_url": reverse("download_error_excel")
            })

        # --- ✅ Valid Excel ---
        file_stream = io.BytesIO()
        df.to_excel(file_stream, index=False)
        request.session['validated_file'] = file_stream.getvalue().decode('latin1')

        return JsonResponse({
            "valid": True,
            "message": f"✅ Link Excel file is valid ({len(df)} records)",
            "count": len(df)
        })

    return JsonResponse({"valid": False, "message": " No file uploaded"})

@csrf_exempt
def validate_map_txt(request):
    if request.method == "POST" and request.FILES.get("map_txt"):
        file = request.FILES["map_txt"]

        try:
            import csv
            from shapely import wkt
            csv.field_size_limit(sys.maxsize) 
            link_data = []

            reader = csv.DictReader(
                (line.decode("utf-8") for line in file),
                delimiter=";"
            )

            for row in reader:
                linkno = str(row["LinkId"]).strip()
                line_wkt = row["Line"].strip()

                # Validate WKT
                try:
                    geom = wkt.loads(line_wkt)
                    if geom.geom_type != "LineString":
                        return JsonResponse({"valid": False, "message": f"Invalid geometry type for LinkId {linkno}"})
                except Exception as e:
                    return JsonResponse({"valid": False, "message": f"Invalid WKT for LinkId {linkno}: {e}"})

                link_data.append((linkno, line_wkt))

            if not link_data:
                return JsonResponse({"valid": False, "message": "No valid link data found in TXT"})

            #   Compare with Excel linkCodes from session
            matched_count = 0
            file_content = request.session.get("validated_file")
            if file_content:
                import io, pandas as pd
                file_stream = io.BytesIO(file_content.encode('latin1'))
                df = pd.read_excel(file_stream)

                excel_linkcodes = set(df["Link_Code"].astype(str))
                matched_count = sum(1 for linkno, _ in link_data if linkno in excel_linkcodes)

            #   Store validated TXT in session
            request.session["validated_txt"] = json.dumps(link_data)

            return JsonResponse({
                "valid": True,
                "message": f"Alignment TXT file is valid   ({len(link_data)} records, {matched_count} matched with Excel)",
                "count": matched_count
            })

        except Exception as e:
            return JsonResponse({"valid": False, "message": f"Error reading TXT file: {e}"})

    return JsonResponse({"valid": False, "message": "No file uploaded"})


from django.http import HttpResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import io

def download_error_excel(request):
    error_excel_b64 = request.session.get("error_excel")
    if not error_excel_b64:
        return HttpResponse("No error Excel available", status=404)

    error_excel = base64.b64decode(error_excel_b64)

    response = HttpResponse(
        error_excel,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = 'attachment; filename="ExcelErrors.xlsx"'
    return response

# Serve empty schema template
def download_template_excel(request):
    required_cols = [
        "Adm_Code", "Link_No", "Link_Code", "Link_Name",
        "Link_Length_Official", "Link_Length_Actual", "Status"
    ]
    wb = Workbook()
    ws = wb.active
    ws.append(required_cols)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output,
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="excel_template.xlsx"'
    return response



@csrf_exempt
def validate_db_file(request):
    
    if request.method == "POST" and request.FILES.get("db_file"):
        file = request.FILES["db_file"]
        admCode = request.POST.get("admcode")
             
        if not admCode:
            return JsonResponse({
                 "valid": False,
                "message": "Please Select Status and Province/Kabupaten"
            })
        # Check if file is .accdb
        if not file.name.lower().endswith('.accdb'):
            return JsonResponse({
                "valid": False,
                "message": "Please upload a Microsoft Access database file (.accdb)"
            })

        # Check if user wants to force download the Excel file
        force_download = request.POST.get("force_download", "false").lower() == "true"

        temp_file_path = None
        try:
            

            # Save uploaded file temporarily
            with tempfile.NamedTemporaryFile(delete=False, suffix='.accdb') as temp_file:
                for chunk in file.chunks():
                    temp_file.write(chunk)
                temp_file_path = temp_file.name
                print(f"Temporary file created: {temp_file_path}")

            if temp_file_path and os.path.exists(temp_file_path):
                # Run validation
                validation_result = runValidationScript(temp_file_path , admCode)

                if validation_result and validation_result.get("success"):
                    # Collect validation summary
                    summary = validation_result.get("summary", {})
                    total_errors = sum(summary.values()) if summary else 0
                    validation_passed = total_errors == 16  # adjust logic if needed

                    excel_file_path = validation_result.get("output_file")

                    # Case 1: Errors exist OR force download
                    if total_errors > 16 or force_download:
                        if excel_file_path and os.path.exists(excel_file_path):
                            from django.http import FileResponse
                            import mimetypes

                            mime_type, _ = mimetypes.guess_type(excel_file_path)
                            if mime_type is None:
                                mime_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

                            file_handle = open(excel_file_path, 'rb')
                            response = FileResponse(file_handle, content_type=mime_type)

                            filename = "validation_errors.xlsx" if total_errors > 16 else "validation_report.xlsx"
                            response['Content-Disposition'] = f'attachment; filename="{filename}"'
                            response['X-Validation-Errors'] = str(total_errors)
                            response['X-Validation-Summary'] = json.dumps(summary)
                            response['X-Validation-Passed'] = str(validation_passed).lower()
                            return response

                        return JsonResponse({
                            "valid": False,
                            "message": "Validation completed but Excel output not found.",
                            "total_errors": total_errors,
                            "summary": summary
                        })

                    # Case 2: No validation errors →  normal response
                    return JsonResponse({
                        "valid": True,
                        "message": "Database validation completed successfully! No validation errors found.",
                        "summary": summary,
                        "total_errors": total_errors,
                        "validation_passed": validation_passed
                    })

                else:
                    # Validation failed
                    error_message = "Database validation failed. Please check the logs."
                    if validation_result and not validation_result.get("success"):
                        error_message = validation_result.get("message", error_message)

                    return JsonResponse({
                        "valid": False,
                        "message": error_message,
                        "validation_result": validation_result
                    })

        except Exception as e:
            print('error:', e)
            return JsonResponse({
                "valid": False,
                "message": f"Error processing Access database: {str(e)}"
            })
        finally:
            # Cleanup temporary file
            if 'temp_file_path' in locals() and temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)
                    print(f"Temporary file cleaned up: {temp_file_path}")
                except Exception as cleanup_error:
                    print('clean up error:', cleanup_error)
                    print(f"Warning: Could not clean up temporary file {temp_file_path}: {cleanup_error}")

    return JsonResponse({"valid": False, "message": "No file uploaded"})

def get_validation_summary(excel_file_path):
    """
    Helper function to get a summary of validation errors from the Excel file
    """
    try:
        import pandas as pd
        excel_file = pd.ExcelFile(excel_file_path)
        summary = {}
        total_errors = 0
        
        for sheet_name in excel_file.sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)
            if not df.empty:
                # Count rows that are not success messages
                error_rows = df[
                    (~df['Record_No'].astype(str).str.contains('NO_ERRORS', na=False)) &
                    (~df['Record_No'].astype(str).str.contains('SUCCESS', na=False)) &
                    (~df['Record_No'].astype(str).str.contains('EMPTY_TABLE', na=False))
                ]
                error_count = len(error_rows)
                summary[sheet_name] = error_count
                total_errors += error_count
        
        return summary, total_errors
    except Exception as e:
        print(f"Error getting validation summary: {e}")
        return {}, 0